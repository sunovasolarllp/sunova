import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice Template"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="0D1321")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    fill_accent = PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    border_total = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # Title Block
    ws.merge_cells("A1:E1")
    ws["A1"] = "SUNOVA SOLAR LLP - PROFORMA INVOICE TEMPLATE"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25
    
    ws["A2"] = "Empaneled Solar rooftop Installer | Opp. KSEB Section, Alappuzha"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells("A2:E2")
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # Customer Details Block
    ws["A4"] = "Consumer Name:"
    ws["A4"].font = font_bold
    ws["B4"] = "[Enter Customer Name]"
    ws["B4"].font = font_normal
    
    ws["A5"] = "Consumer No:"
    ws["A5"].font = font_bold
    ws["B5"] = "[Enter 13-digit KSEB Consumer No]"
    ws["B5"].font = font_normal
    
    ws["D4"] = "Proforma Qtn No:"
    ws["D4"].font = font_bold
    ws["E4"] = "SNS/PRO/2026/XXXX"
    
    ws["D5"] = "Date:"
    ws["D5"].font = font_bold
    ws["E5"] = "=TODAY()"
    
    # Table Headers
    headers = ["Item No", "Component Description & Specification", "Quantity", "Unit Rate (₹)", "Total Value (₹)"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws.row_dimensions[7].height = 20
    
    # Table Content
    items = [
        (1, "MNRE ALMM Approved Mono PERC Solar PV Modules (540Wp+)", 6, 16000),
        (2, "On-Grid Tied Solar Inverter with Anti-islanding & Wi-Fi logger", 1, 38000),
        (3, "Hot-dip Galvanized Iron (GI) high-rise structure with anchoring", 1, 18000),
        (4, "ACDB/DCDB protection boxes, SPDs, Copper Cables & earthings", 1, 14000),
        (5, "Liaisoning, net-metering feasibility processing & commissioning", 1, 12000),
    ]
    
    for row_idx, item in enumerate(items, 8):
        ws.cell(row=row_idx, column=1, value=item[0]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=item[1])
        ws.cell(row=row_idx, column=3, value=item[2]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=item[3])
        ws.cell(row=row_idx, column=5, value=f"=C{row_idx}*D{row_idx}") # Formula for total
        
        for col_idx in range(1, 6):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = font_normal
            c.border = border_thin
            if col_idx in [4, 5]:
                c.number_format = '₹#,##0'
    
    # Summary Calculations
    # Subtotal
    ws.cell(row=13, column=2, value="Subtotal Basic Cost").font = font_bold
    ws.cell(row=13, column=5, value="=SUM(E8:E12)").font = font_bold
    ws.cell(row=13, column=5).number_format = '₹#,##0'
    ws.cell(row=13, column=5).border = Border(top=Side(style='thin', color='000000'))
    
    # GST tax
    ws.cell(row=14, column=2, value="GST Tax @ 13.8%").font = font_normal
    ws.cell(row=14, column=5, value="=E13*0.138").font = font_normal
    ws.cell(row=14, column=5).number_format = '₹#,##0'
    
    # Grand Total
    ws.cell(row=15, column=2, value="Grand Total Contract Value").font = font_bold
    ws.cell(row=15, column=5, value="=E13+E14").font = font_bold
    ws.cell(row=15, column=5).number_format = '₹#,##0'
    ws.cell(row=15, column=5).border = border_total
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    # Ensure assets dir exists
    os.makedirs("assets", exist_ok=True)
    
    wb.save("assets/proforma-invoice-template.xlsx")
    print("Template created successfully!")

if __name__ == "__main__":
    create_template()
